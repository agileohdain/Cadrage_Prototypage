#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make-matrice-template.py — génère le modèle VIERGE Matrice.xlsx (cadrage).

Produit un classeur à 2 feuilles :
  * KPI     — 11 colonnes (Thèmes … Commentaires), prêt à remplir.
  * Filtres — listes de référence (Priorités, Fiabilités).

Listes déroulantes (Data Validation) posées sur :
  * Priorités             (col E) -> Haute | Moyenne | Basse
  * Fiabilités des données (col H) -> Fiable | Partielle | Inconnue

Usage :
  python make-matrice-template.py [output.xlsx]

Sans argument : écrit Matrice.xlsx à la racine du dépôt.
"""
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("ERREUR: openpyxl manquant -> pip install openpyxl\n")
    sys.exit(1)

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", ".."))

HEADERS = [
    "Thèmes",
    "Sujets",
    "Indicateurs",
    "Descriptions",
    "Priorités",
    "Provenance systèmes sources",
    "Sources des données",
    "Fiabilités des données",
    "Axes d'analyses",
    "Formules",
    "Commentaires",
]

PRIORITES = ["Haute", "Moyenne", "Basse"]
FIABILITES = ["Fiable", "Partielle", "Inconnue"]

# Colonnes (1-indexées) concernées par les listes déroulantes.
COL_PRIORITE = HEADERS.index("Priorités") + 1            # E
COL_FIABILITE = HEADERS.index("Fiabilités des données") + 1  # H
N_ROWS = 200  # lignes pré-équippées de validation


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = fill
        c.font = font
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 30


def set_widths(ws):
    widths = {
        "Thèmes": 18, "Sujets": 22, "Indicateurs": 28, "Descriptions": 40,
        "Priorités": 12, "Provenance systèmes sources": 24,
        "Sources des données": 28, "Fiabilités des données": 16,
        "Axes d'analyses": 30, "Formules": 26, "Commentaires": 30,
    }
    for ci, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 18)


def build_kpi_sheet(wb):
    ws = wb.active
    ws.title = "KPI"
    ws.append(HEADERS)
    style_header(ws)
    set_widths(ws)
    ws.freeze_panes = "A2"

    # Bordures légères + alignement haut sur les lignes saisissables.
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(vertical="top", wrap_text=True)
    for r in range(2, N_ROWS + 2):
        for ci in range(1, len(HEADERS) + 1):
            c = ws.cell(row=r, column=ci)
            c.border = border
            c.alignment = wrap_top

    # Listes déroulantes (Data Validation) — formule inline (portabilité max).
    dv_prio = DataValidation(
        type="list", formula1='"%s"' % ",".join(PRIORITES),
        allow_blank=True, showDropDown=False)
    dv_prio.error = "Choisir : Haute, Moyenne ou Basse."
    dv_prio.errorTitle = "Priorité invalide"
    dv_prio.add("%s2:%s%d" % (get_column_letter(COL_PRIORITE),
                              get_column_letter(COL_PRIORITE), N_ROWS + 1))
    ws.add_data_validation(dv_prio)

    dv_fiab = DataValidation(
        type="list", formula1='"%s"' % ",".join(FIABILITES),
        allow_blank=True, showDropDown=False)
    dv_fiab.error = "Choisir : Fiable, Partielle ou Inconnue."
    dv_fiab.errorTitle = "Fiabilité invalide"
    dv_fiab.add("%s2:%s%d" % (get_column_letter(COL_FIABILITE),
                              get_column_letter(COL_FIABILITE), N_ROWS + 1))
    ws.add_data_validation(dv_fiab)


def build_filtres_sheet(wb):
    ws = wb.create_sheet("Filtres")
    ws.append(["Priorités"])
    for v in PRIORITES:
        ws.append([v])
    ws.append([])
    ws.append(["Fiabilités"])
    for v in FIABILITES:
        ws.append([v])
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for r in (1, len(PRIORITES) + 3):
        c = ws.cell(row=r, column=1)
        c.fill = fill
        c.font = font
    ws.column_dimensions["A"].width = 16


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "Matrice.xlsx")
    wb = openpyxl.Workbook()
    build_kpi_sheet(wb)
    build_filtres_sheet(wb)
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    sys.stderr.write("OK: modèle généré -> %s\n" % out)


if __name__ == "__main__":
    main()
