#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
parse-matrice.py — lit le Matrice.xlsx rempli par le conseiller + client et
produit cadrage.json (structure normalisée consommée par le skill
cadrage-prototypage puis generate-cadrage.py).

Feuille KPI (11 colonnes) :
  Thèmes | Sujets | Indicateurs | Descriptions | Priorités |
  Provenance systèmes sources | Sources des données | Fiabilités des données |
  Axes d'analyses | Formules | Commentaires

Normalisation :
  * Priorités  -> Haute | Moyenne | Basse
  * Fiabilités -> fiable | partielle | inconnue  (3 niveaux)
    - cellule vide + « Sources des données » vide -> inconnue (source non identifiée)
    - « non fiable / incertain / à valider / estimé … » -> partielle
    - « fiable / validé / sûr » -> fiable

Usage :
  python parse-matrice.py <matrice.xlsx> [-o cadrage.json]
  Défaut : écrit cadrage.json à côté du classeur.
"""
import os
import re
import sys
import json
import unicodedata

try:
    import openpyxl
except ImportError:
    sys.stderr.write("ERREUR: openpyxl manquant -> pip install openpyxl\n")
    sys.exit(1)

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# Correspondance clé normalisée -> champ cadrage.json
COL_MAP = {
    "themes": "theme",
    "sujets": "sujet",
    "indicateurs": "indicateur",
    "descriptions": "description",
    "priorites": "priorite",
    "provenance systemes sources": "systeme_source",
    "sources des donnees": "source_donnees",
    "fiabilites des donnees": "fiabilite",
    "axes d analyses": "axes",
    "axes danalyses": "axes",
    "formules": "formule",
    "commentaires": "commentaires",
}

PRIO_BUCKETS = ("Haute", "Moyenne", "Basse")
FIAB_BUCKETS = ("fiable", "partielle", "inconnue")


def norm(s):
    """lowercase, sans accents, non-alphanum -> espace, collapsé."""
    if s is None:
        return ""
    s = str(s).lower().strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s


def norm_prio(v):
    n = norm(v)
    if not n:
        return None
    if n.startswith("haut") or "haute" in n:
        return "Haute"
    if n.startswith("moy") or "moyen" in n:
        return "Moyenne"
    if n.startswith("bas") or "basse" in n:
        return "Basse"
    return None


def norm_fiab(v, has_source):
    n = norm(v)
    if not n:
        return "inconnue" if not has_source else "fiable"
    if "inconnu" in n:
        return "inconnue"
    if "non fiable" in n or "nonfiab" in n:
        return "partielle"
    if "partiel" in n:
        return "partielle"
    if any(k in n for k in (
            "incertain", "a valider", "avalider", "estime", "provisoire",
            "non valide", "nonvalide", "approximatif", "aleatoire", "flou")):
        return "partielle"
    if any(k in n for k in ("fiable", "valide", "sure", "sur ", "fiable")):
        return "fiable"
    return "partielle" if has_source else "inconnue"


def split_axes(v):
    if not v:
        return []
    parts = re.split(r"[\n;|/]+", str(v))
    parts = [re.sub(r"\s*[,]?\s*$", "", p).strip(" -—\t") for p in parts]
    return [p for p in parts if p]


def detect_columns(header_row):
    """Retourne {champ: index_col} d'après la ligne d'en-têtes (robuste aux accents)."""
    cols = {}
    for ci, raw in enumerate(header_row):
        key = COL_MAP.get(norm(raw))
        if key:
            cols[key] = ci
    return cols


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "KPI" not in wb.sheetnames:
        raise SystemExit("ERREUR: feuille 'KPI' introuvable dans %s" % path)
    ws = wb["KPI"]

    rows = list(ws.iter_rows(values_only=True))
    # trouve la ligne d'en-têtes (la 1re contenant 'Indicateurs')
    header_idx = None
    for i, r in enumerate(rows):
        if any(norm(c) == "indicateurs" for c in r if c is not None):
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit("ERREUR: en-tête 'Indicateurs' introuvable dans la feuille KPI.")
    cols = detect_columns(rows[header_idx])
    required = ("indicateur", "priorite", "fiabilite")
    for req in required:
        if req not in cols:
            raise SystemExit("ERREUR: colonne requise manquante (%s) dans la feuille KPI." % req)

    themes = []
    theme_index = {}
    stats = {"total": 0, "Haute": 0, "Moyenne": 0, "Basse": 0, "sans_prio": 0,
             "fiable": 0, "partielle": 0, "inconnue": 0}
    warnings = []
    kid = 0

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        indic = (r[cols["indicateur"]] if "indicateur" in cols else None)
        if indic is None or str(indic).strip() == "":
            continue  # ligne vide / sans KPI

        kid += 1
        src_donnees = str(r[cols["source_donnees"]]).strip() if "source_donnees" in cols and r[cols["source_donnees"]] else ""
        has_source = bool(src_donnees)

        prio_raw = r[cols["priorite"]] if "priorite" in cols and r[cols["priorite"]] else ""
        prio = norm_prio(prio_raw)
        if prio is None and str(prio_raw).strip():
            warnings.append("K%02d priorité non reconnue '%s' -> classée sans_prio" % (kid, prio_raw))

        fiab_raw = r[cols["fiabilite"]] if "fiabilite" in cols and r[cols["fiabilite"]] else ""
        fiab = norm_fiab(fiab_raw, has_source)

        # arbre Thèmes > Sujets
        theme = str(r[cols["theme"]]).strip() if "theme" in cols and r[cols["theme"]] else "(Sans thème)"
        sujet = str(r[cols["sujet"]]).strip() if "sujet" in cols and r[cols["sujet"]] else "(Sans sujet)"

        if theme not in theme_index:
            theme_index[theme] = len(themes)
            themes.append({"theme": theme, "sujets": []})
        tnode = themes[theme_index[theme]]
        snode = next((s for s in tnode["sujets"] if s["sujet"] == sujet), None)
        if snode is None:
            snode = {"sujet": sujet, "kpis": []}
            tnode["sujets"].append(snode)

        kpi = {
            "id": "K%02d" % kid,
            "indicateur": str(indic).strip(),
            "description": str(r[cols["description"]]).strip() if "description" in cols and r[cols["description"]] else "",
            "priorite": prio or "",
            "systeme_source": str(r[cols["systeme_source"]]).strip() if "systeme_source" in cols and r[cols["systeme_source"]] else "",
            "source_donnees": src_donnees,
            "fiabilite": fiab,
            "axes": split_axes(r[cols["axes"]] if "axes" in cols and r[cols["axes"]] else ""),
            "formule": str(r[cols["formule"]]).strip() if "formule" in cols and r[cols["formule"]] else "",
            "commentaires": str(r[cols["commentaires"]]).strip() if "commentaires" in cols and r[cols["commentaires"]] else "",
        }
        snode["kpis"].append(kpi)

        # stats
        stats["total"] += 1
        if prio in PRIO_BUCKETS:
            stats[prio] += 1
        else:
            stats["sans_prio"] += 1
        stats[fiab] += 1

    if stats["total"] == 0:
        raise SystemExit("ERREUR: aucun indicateur trouvé — le Matrice.xlsx est vide. "
                         "Remplissez la feuille KPI avant de relancer.")

    # compteurs cumulés par périmètre priorité (pour la question guidée)
    h, m, b = stats["Haute"], stats["Moyenne"], stats["Basse"]
    stats["scope"] = {
        "haute": h,
        "haute_moyenne": h + m,
        "haute_moyenne_basse": h + m + b,
    }
    return {"source": os.path.basename(path), "stats": stats, "warnings": warnings,
            "themes": themes}


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("usage: python parse-matrice.py <matrice.xlsx> [-o cadrage.json]\n")
        sys.exit(2)
    path = sys.argv[1]
    if not os.path.exists(path):
        raise SystemExit("ERREUR: fichier introuvable: %s" % path)
    out = None
    if "-o" in sys.argv:
        out = sys.argv[sys.argv.index("-o") + 1]
    if not out:
        out = os.path.join(os.path.dirname(os.path.abspath(path)), "cadrage.json")

    cadrage = parse(path)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(cadrage, f, ensure_ascii=False, indent=2)
        f.write("\n")

    s = cadrage["stats"]
    sys.stderr.write("OK: %d indicateurs (%d thèmes) -> %s\n"
                     % (s["total"], len(cadrage["themes"]), out))
    sys.stderr.write("    Priorités : Haute=%d  Moyenne=%d  Basse=%d  (sans=%d)\n"
                     % (s["Haute"], s["Moyenne"], s["Basse"], s["sans_prio"]))
    sys.stderr.write("    Fiabilité : fiable=%d  partielle=%d  inconnue=%d\n"
                     % (s["fiable"], s["partielle"], s["inconnue"]))
    for w in cadrage["warnings"]:
        sys.stderr.write("    ATTENTION: %s\n" % w)


if __name__ == "__main__":
    main()
